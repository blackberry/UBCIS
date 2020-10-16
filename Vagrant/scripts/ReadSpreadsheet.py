################################################################################
# Name   : ReadSpreadsheet - Read excel spreadsheets
# Author : Alexander Parent
#
# Copyright 2020 BlackBerry Limited
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
#    Unless required by applicable law or agreed to in writing, software
#    distributed under the License is distributed on an "AS IS" BASIS,
#    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#    See the License for the specific language governing permissions and
#    limitations under the License.
################################################################################

import openpyxl
import json

config = None
with open('/config/config.json', 'r') as json_file:
    config = json.load(json_file)

output = "/results/output/" + config['output']
print('Processing {0}...'.format(output))
loadedJson = None

legitimateRatings = ['tp', 'c', 'mm', 'd', 'fp']
incompleteImages = []

with open(output + ".json", 'r') as json_file:
    loadedJson = json.load(json_file)

    wb = openpyxl.load_workbook(filename=output + ".xlsx")

    for worksheet in wb.worksheets:
        realName = worksheet.title
        print('Processing {0}...'.format(realName))
        imageVulns = loadedJson['images'][realName]

        row = 2
        cell = worksheet.cell(2, 2).value

        while cell != '' and cell is not None:
            trueVuln = worksheet.cell(row, 7).value
            notes = worksheet.cell(row, 8).value

            filtered = list(filter(lambda x: x['vulnerability'] == cell, imageVulns))
            if len(filtered) == 1:
                filtered[0]['trueVulnerability'] = trueVuln

                if (trueVuln is None or trueVuln.lower() not in legitimateRatings) and realName not in incompleteImages:
                    incompleteImages.append(realName)

                filtered[0]['notes'] = notes
            else:
                print(cell, trueVuln, notes)

            row += 1
            cell = worksheet.cell(row, 2).value

with open(output + ".json", 'w') as json_file:
    json.dump(loadedJson, json_file)

imageKeys = list(loadedJson['images'].keys())
for image in imageKeys:
    if image in incompleteImages:
        print('Removing image without complete vulnerability marks from benchmark', image)
        del loadedJson['images'][image]
    else:
        for vuln in loadedJson['images'][image]:
            del vuln['parsers']
            vuln['version'] = ''
            vuln['notes'] = ''

loadedJson['parsers'] = []

with open("/benchmark/benchmark.json", 'w') as json_file:
    json.dump(loadedJson, json_file)
